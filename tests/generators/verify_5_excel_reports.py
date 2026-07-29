import os
import openpyxl

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")

target_files = [
    ("Selenium_Test_Report.xlsx", "Selenium Test Cases"),
    ("Unit_Test_Report.xlsx", "Unit Test Cases"),
    ("Load_Test_Report.xlsx", "Load Test Cases"),
    ("Vulnerability_Test_Report.xlsx", "Vulnerability Test Cases"),
    ("Validation_Test_Report.xlsx", "Validation Test Cases"),
]

print("=" * 60)
print("STARTING AUDIT AND VERIFICATION OF 5 EXCEL TEST REPORTS")
print("=" * 60)

all_passed = True

for file_name, sheet_name in target_files:
    file_path = os.path.join(REPORTS_DIR, file_name)
    if not os.path.exists(file_path):
        print(f"[FAIL] File missing: {file_name}")
        all_passed = False
        continue

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"[FAIL] Sheet '{sheet_name}' not found in {file_name}")
        all_passed = False
        continue

    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    header_row = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    status_col_idx = len(header_row)

    data_rows = max_row - 1
    passed_count = 0
    failed_count = 0

    for r in range(2, max_row + 1):
        status_val = str(ws.cell(row=r, column=status_col_idx).value or '').strip()
        if status_val.upper() == "PASS":
            passed_count += 1
        else:
            failed_count += 1

    print(f"File: {file_name}")
    print(f"  - Target Sheet: '{sheet_name}'")
    print(f"  - Total Rows: {max_row} (1 Header + {data_rows} Data Rows)")
    print(f"  - Passed Test Cases: {passed_count}")
    print(f"  - Failed Test Cases: {failed_count}")

    if max_row == 301 and data_rows == 300 and passed_count == 300 and failed_count == 0:
        print(f"  - VERIFICATION AUDIT: [PASS] (Exact 301 total lines, 300 test cases, 100% PASS)")
    else:
        print(f"  - VERIFICATION AUDIT: [FAIL]")
        all_passed = False
    print("-" * 60)

if all_passed:
    print("=" * 60)
    print("ALL 5 EXCEL REPORTS PASSED VERIFICATION AUDIT WITH 100% SUCCESS!")
    print("=" * 60)
else:
    print("AUDIT FAILED FOR ONE OR MORE FILES.")
